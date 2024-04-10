from openpyxl import load_workbook

wb = load_workbook('ex_2.xlsx')
sh1 = wb['Sheet1']

rows = sh1.max_row
columns = sh1.max_column

# for r in range(1,2):
#     for c in range(1,columns+1):
#         print(sh1.cell(row = r,column = c).value)



# sh1['H1'] = "Employee Salary"
payroll_data ={}
with open('work/payroll.txt','r') as f:
    for line in f:
        emp_code,work_days = line.strip().split(',')
        salary = int(work_days)*1000
        payroll_data[emp_code] = salary

for i, value in enumerate(payroll_data.values(), start=2):
    cell = sh1.cell(row=i, column=8)
    cell.value = value

wb.save('ex_2.xlsx')





# =======================================================================================================
# import os

# # Specify the folder name
# folder_name = "new_folder"

# # Specify the file name
# file_name = "new_file.txt"

# # Combine the current directory path with the folder name
# folder_path = os.path.join(os.getcwd(), folder_name)

# # Create the folder if it doesn't exist
# if not os.path.exists(folder_path):
#     os.makedirs(folder_path)

# # Combine the folder path and file name to create the full path for the file
# file_path = os.path.join(folder_path, file_name)

# # Open a file in write mode in the specified folder
# with open(file_path, "w") as file:
#     # Write some content to the file
#     file.write("Hello, this is the content of the file in the new folder.")

# =======================================================================================================
# import os

# folder_path = "Week-2\\8_April"
# file_name = "filename.txt"


# # # Check if the folder exists, if not create it
# # if not os.path.exists(folder_path):
# #     os.makedirs(folder_path)


# # Combine the folder path and file name
# file_path = os.path.join(folder_path, file_name)


# # Open a file in write mode in the specified folder

# with open(file_path, "w") as file:
#     file.write("Hello, this is the content of the file.")
