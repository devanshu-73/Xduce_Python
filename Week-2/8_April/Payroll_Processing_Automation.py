import os
import datetime
import shutil
from openpyxl import load_workbook

# Function to check if current date is past 20th of the month
def past_20th():
    today = datetime.datetime.now()
    return today.day <= 20

# Function to check if payroll file is empty
def payroll_empty(file_path):
    return os.path.exists(file_path) and os.path.getsize(file_path) == 0

# Function to copy file from source folder to destination folder
def copy_file(source_path, dest_folder):
    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)

    file_name = os.path.basename(source_path)
    dest_path = os.path.join(dest_folder, file_name)

    with open(source_path, 'r') as src, open(dest_path, 'w') as dst:
        dst.write(src.read())

    print(f"File copied to {dest_folder}.")

    return dest_path

# Function to process payroll and generate Excel file
def process_payroll(payroll_file_path, employee_data_path):
    # print("==============================",payroll_file_path, employee_data_path)

    # Read payroll data and calculate salary
    payroll_data = {}
 
    with open(payroll_file_path, 'r') as file:
        
        for line in file:

            emp_code, work_days = line.strip().split(',') # print("Employee code:", emp_code,"|| Work days :", work_days)
            salary = int(work_days) * 1000
            payroll_data[emp_code] = salary

    # Read employee data and create a workbook
    wb = load_workbook(employee_data_path)
    sh1 = wb['Sheet1']
    rows = sh1.max_row
    columns = sh1.max_column 

    for i, value in enumerate(payroll_data.values(), start=2):
        cell = sh1.cell(row=i, column=8)
        cell.value = value

    wb.save(employee_data_path)

# Main function
def main():
    if not past_20th():
        print("Payroll processing is scheduled after the 20th of the month.")
        return

    payroll_file_path = "account/payroll.txt"
    employee_data_path = "ex_2.xlsx"

    if payroll_empty(payroll_file_path):
        print("Payroll file is empty. No processing needed.")

    work_folder = "work/"
    new_file_path = copy_file(payroll_file_path, work_folder)

    output_file_path = process_payroll(new_file_path, employee_data_path)

if __name__ == "__main__":
    main()

